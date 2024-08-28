import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell
from typing import List, Any


class ExcelTableExtractor:
    def __init__(self, file_path: str):
        """
        Initialize the ExcelTableExtractor with the given file path.

        :param file_path: Path to the Excel file.
        """
        self.file_path = file_path
        self.wb = load_workbook(file_path, data_only=True)
        self.sheet = self.wb.active
        self.tables = []

    def has_border(self, cell: Cell) -> bool:
        """
        Check if a cell has any border.

        :param cell: The cell to check.
        :return: True if the cell has any border, False otherwise.
        """
        return self.has_row_border(cell) or self.has_col_border(cell)

    def has_row_border(self, cell: Cell) -> bool:
        """
        Check if a cell has a top or bottom border.

        :param cell: The cell to check.
        :return: True if the cell has a top or bottom border, False otherwise.
        """
        border = cell.border
        return bool(border.top.style or border.bottom.style)

    def has_col_border(self, cell: Cell) -> bool:
        """
        Check if a cell has a left or right border.

        :param cell: The cell to check.
        :return: True if the cell has a left or right border, False otherwise.
        """
        border = cell.border
        return bool(border.left.style or border.right.style)

    def extract_tables(self):
        """
        Extract tables from the Excel sheet based on cell borders.
        """
        rows = list(self.sheet.iter_rows())
        flags = np.zeros((len(rows), len(rows[0])))

        def get_table(start_row_idx: int, start_col_idx: int) -> List[List[Any]]:
            header_row = rows[start_row_idx]
            end_col_idx = len(header_row) - 1
            for col_idx in range(start_col_idx, len(header_row)):
                cell = header_row[col_idx]
                if not self.has_border(cell):
                    end_col_idx = col_idx - 1
                    break

            table = []
            for row_idx in range(start_row_idx, len(rows)):
                cell = rows[row_idx][start_col_idx]
                if not self.has_border(cell):
                    break
                row_data = []
                for col_idx in range(start_col_idx, end_col_idx + 1):
                    row_data.append(rows[row_idx][col_idx].value)
                    flags[row_idx][col_idx] = 1
                table.append(row_data)
            return table

        for row_idx, row in enumerate(rows):
            for col_idx, col in enumerate(row):
                if flags[row_idx][col_idx] == 0 and self.has_border(col):
                    table = get_table(row_idx, col_idx)
                    self.tables.append(table)

    def to_dataframes(self) -> List[pd.DataFrame]:
        """
        Convert extracted tables to a list of pandas DataFrames.

        :return: List of DataFrames.
        """
        return [pd.DataFrame(table) for table in self.tables]

    def display_tables(self):
        """
        Display the extracted tables as pandas DataFrames.
        """
        dataframes = self.to_dataframes()
        for i, df in enumerate(dataframes):
            print(f"Table {i + 1}:")
            print(df)
            print("\n")
